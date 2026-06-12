"""
Fabric Schedule Extractor
==========================
Connects to Microsoft Fabric and Power BI, lists all workspaces the signed-in
user can access, lets the user choose one by name, then extracts all schedule
information for Data Pipelines, Semantic Models, Dataflows, Notebooks, and
Spark Job Definitions in that workspace.

Requirements:
  - Python 3.8 or later
  - Internet connection
  - A Microsoft account with access to the target Fabric workspace
  - config.py in the repo root with your CLIENT_ID (copy from config.example.py)

First-time setup:
  Run the script. Missing packages are installed automatically and a browser
  window opens for Microsoft sign-in. The token is cached so subsequent runs
  do not require signing in again.

Cache / stale data:
  A checkpoint file is saved per workspace so a long run can be resumed if
  interrupted. The checkpoint is workspace-specific and automatically discarded
  if it is more than 1 hour old, ensuring you always get fresh data on a new
  session. You can also choose to start fresh when prompted.

Output:
  An Excel file saved to your Downloads folder named after the workspace, with
  one sheet per selected item type plus an Active Schedules summary sheet and
  a Pipeline Relationships sheet (if pipelines were included).
"""

# Author: Abigail Woolley, AmaliTech
# Built with AI assistance

import sys
import subprocess
from pathlib import Path

# ── Auto-install required packages ───────────────────────────────────────────
REQUIRED = ['msal', 'requests', 'openpyxl', 'rich']
for pkg in REQUIRED:
    try:
        __import__(pkg)
    except ImportError:
        print(f'Installing {pkg}...')
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg, '--quiet'])
        print(f'  {pkg} installed.')

import msal
import requests
import json
import base64
import time
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.progress import (
    Progress, SpinnerColumn, BarColumn, TextColumn,
    TimeElapsedColumn, TimeRemainingColumn, MofNCompleteColumn,
)
from rich.table import Table
from rich.panel import Panel
from rich import box

sys.stdout.reconfigure(encoding='utf-8')
console = Console()

# ── Load CLIENT_ID from config.py ─────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent.parent))
try:
    from config import CLIENT_ID
except ImportError:
    console.print('[bold red]ERROR: config.py not found.[/bold red]')
    console.print('Copy config.example.py to config.py and add your CLIENT_ID.')
    sys.exit(1)

# ── Configuration ─────────────────────────────────────────────────────────────
AUTHORITY  = 'https://login.microsoftonline.com/organizations'
TIMEOUT    = 30
CACHE_FILE = Path.home() / '.fabric_token_cache.bin'
BASE       = 'https://api.fabric.microsoft.com/v1'
PBI_BASE   = 'https://api.powerbi.com/v1.0/myorg'

# ── Helpers ───────────────────────────────────────────────────────────────────
def trunc(s, n=52):
    return (s[:n - 3] + '...') if len(s) > n else s


def make_progress():
    return Progress(
        SpinnerColumn(style='bold cyan'),
        TextColumn('[dim]{task.fields[item]:<54}', justify='left'),
        BarColumn(bar_width=30, style='cyan', complete_style='bold cyan'),
        MofNCompleteColumn(),
        TimeElapsedColumn(),
        TextColumn('[dim]ETA'),
        TimeRemainingColumn(),
        console=console,
        transient=True,
    )


# ── Banner ────────────────────────────────────────────────────────────────────
console.print()
console.print(Panel.fit(
    '[bold cyan]Fabric Schedule Extractor[/bold cyan]',
    border_style='cyan',
))
console.print()

# ── Authentication ─────────────────────────────────────────────────────────────
FAB_SCOPES = ['https://api.fabric.microsoft.com/.default']
PBI_SCOPES = ['https://analysis.windows.net/powerbi/api/.default']

cache = msal.SerializableTokenCache()
if CACHE_FILE.exists():
    cache.deserialize(CACHE_FILE.read_text(encoding='utf-8'))

app = msal.PublicClientApplication(CLIENT_ID, authority=AUTHORITY, token_cache=cache)


def get_token(scopes):
    accounts = app.get_accounts()
    token    = None
    if accounts:
        token = app.acquire_token_silent(scopes, account=accounts[0])
    if not token:
        console.print('[yellow]Opening browser for Microsoft sign-in...[/yellow]')
        try:
            token = app.acquire_token_interactive(scopes)
        except Exception:
            flow = app.initiate_device_flow(scopes)
            console.print(flow['message'])
            token = app.acquire_token_by_device_flow(flow)
    if 'access_token' not in token:
        console.print('[bold red]ERROR: Authentication failed.[/bold red]')
        console.print(token.get('error_description', 'Unknown error'))
        sys.exit(1)
    if cache.has_state_changed:
        CACHE_FILE.write_text(cache.serialize(), encoding='utf-8')
    return token['access_token']


console.print('[dim]Authenticating...[/dim]')
fab_at = get_token(FAB_SCOPES)
pbi_at = get_token(PBI_SCOPES)
FAB = {'Authorization': 'Bearer ' + fab_at, 'Content-Type': 'application/json'}
PBI = {'Authorization': 'Bearer ' + pbi_at}
console.print('[bold green]✓[/bold green] Authenticated successfully.')
console.print()

run_start     = time.time()
section_times = {}

# ── Workspace Selection ────────────────────────────────────────────────────────
console.print('[dim]Fetching available workspaces...[/dim]')
ws_resp = requests.get(BASE + '/workspaces', headers=FAB, timeout=TIMEOUT)
if ws_resp.status_code != 200:
    console.print(f'[bold red]ERROR: Could not list workspaces ({ws_resp.status_code})[/bold red]')
    console.print(ws_resp.text[:500])
    sys.exit(1)
all_workspaces = ws_resp.json().get('value', [])
console.print(f'[bold green]✓[/bold green] {len(all_workspaces)} workspaces available.')
console.print()

search      = console.input('[bold]Enter workspace name[/bold] [dim](or part of it)[/dim]: ').strip().lower()
search_norm = search.replace(' ', '').replace('-', '')
matches     = [
    w for w in all_workspaces
    if search in w['displayName'].lower()
    or search_norm in w['displayName'].lower().replace(' ', '').replace('-', '')
]

if not matches:
    console.print(f'\n[red]No workspace found matching "[bold]{search}[/bold]". Available workspaces:[/red]')
    for w in sorted(all_workspaces, key=lambda x: x['displayName']):
        console.print(f'  {w["displayName"]}')
    sys.exit(1)

if len(matches) == 1:
    selected_workspace = matches[0]
else:
    console.print('\n[yellow]Multiple matches found:[/yellow]')
    for i, w in enumerate(matches, 1):
        console.print(f'  [[bold]{i}[/bold]] {w["displayName"]}')
    choice = console.input('Enter number: ').strip()
    try:
        selected_workspace = matches[int(choice) - 1]
    except (ValueError, IndexError):
        console.print('[red]Invalid choice.[/red]')
        sys.exit(1)

WORKSPACE_ID   = selected_workspace['id']
WORKSPACE_NAME = selected_workspace['displayName']
console.print()
console.print(f'  [bold]Workspace:[/bold] [cyan]{WORKSPACE_NAME}[/cyan]')
console.print(f'  [bold]ID:[/bold]        [dim]{WORKSPACE_ID}[/dim]')
console.print()

OUTPUT_PATH = Path.home() / 'Downloads' / f'{WORKSPACE_NAME} - Schedule Inventory.xlsx'

# ── Item type selection ────────────────────────────────────────────────────────
ITEM_MENU = [
    ('pipelines',       'Data Pipelines'),
    ('semantic_models', 'Semantic Models'),
    ('dataflows',       'Dataflows'),
    ('notebooks',       'Notebooks'),
    ('spark_jobs',      'Spark Job Definitions'),
]

console.print('[bold]Which item types would you like to extract?[/bold]')
console.print()
for i, (_, label) in enumerate(ITEM_MENU, 1):
    console.print(f'  [[bold cyan]{i}[/bold cyan]] {label}')
console.print(f'  [[bold cyan]A[/bold cyan]] All of the above [dim](default)[/dim]')
console.print()
raw = console.input('Enter numbers separated by commas, or press [bold]Enter[/bold] for all: ').strip().upper()

if not raw or raw == 'A':
    selected_types = {key for key, _ in ITEM_MENU}
else:
    selected_types = set()
    for s in raw.split(','):
        s = s.strip()
        if s.isdigit():
            idx = int(s) - 1
            if 0 <= idx < len(ITEM_MENU):
                selected_types.add(ITEM_MENU[idx][0])
    if not selected_types:
        console.print('[yellow]  No valid selection — defaulting to all.[/yellow]')
        selected_types = {key for key, _ in ITEM_MENU}

console.print()
console.print('[bold]Extracting:[/bold]')
for key, label in ITEM_MENU:
    if key in selected_types:
        console.print(f'  [green]+[/green] {label}')
console.print()

include_unscheduled = console.input(
    'Include items with [bold]no schedule[/bold] in the output? [[bold]Y[/bold]/n]: '
).strip().lower() != 'n'
console.print()

# ── Checkpoint ────────────────────────────────────────────────────────────────
CHECKPOINT_FILE = Path.home() / f'.fabric_checkpoint_{WORKSPACE_ID}.json'
checkpoint      = {}

if CHECKPOINT_FILE.exists():
    age_seconds = time.time() - CHECKPOINT_FILE.stat().st_mtime
    age_minutes = int(age_seconds / 60)
    if age_seconds < 3600:
        resp = console.input(
            f'[yellow]Found cached pipeline data from {age_minutes} minute(s) ago.[/yellow] '
            f'Resume from where it left off? [y/[bold]N[/bold]]: '
        ).strip().lower()
        if resp == 'y':
            try:
                checkpoint = json.loads(CHECKPOINT_FILE.read_text(encoding='utf-8'))
                console.print(f'[green]✓[/green] Resuming — {len(checkpoint)} pipelines already cached.')
            except Exception:
                checkpoint = {}
        else:
            CHECKPOINT_FILE.unlink()
            console.print('[dim]  Starting fresh.[/dim]')
    else:
        console.print(f'[dim]  Cached data is {age_minutes} minutes old — discarding and starting fresh.[/dim]')
        CHECKPOINT_FILE.unlink()

console.print()

# ── API helpers ────────────────────────────────────────────────────────────────
def fmt_time(t):
    if not t or t == 'Never':
        return t or 'Never'
    try:
        dt = datetime.fromisoformat(str(t).replace('Z', '+00:00'))
        return dt.strftime('%Y-%m-%d %H:%M UTC')
    except Exception:
        return str(t)


def get_definition_parts(item_id):
    try:
        r = requests.post(
            BASE + '/workspaces/' + WORKSPACE_ID + '/items/' + item_id + '/getDefinition',
            headers=FAB, timeout=TIMEOUT
        )
        if r.status_code == 202:
            op_url = r.headers.get('Location')
            for _ in range(15):
                time.sleep(2)
                try:
                    r2 = requests.get(op_url, headers=FAB, timeout=TIMEOUT)
                    if r2.status_code == 200:
                        return r2.json().get('definition', {}).get('parts', [])
                except Exception:
                    pass
            return []
        return r.json().get('definition', {}).get('parts', []) if r.status_code == 200 else []
    except Exception:
        return None


def parse_schedule(parts):
    for p in parts:
        if p.get('path') == '.schedules':
            try:
                data    = json.loads(base64.b64decode(p['payload']).decode('utf-8', errors='replace'))
                results = []
                for s in data.get('schedules', []):
                    cfg = s.get('configuration', {})
                    results.append({
                        'enabled':    s.get('enabled', False),
                        'type':       cfg.get('type', ''),
                        'times':      ', '.join(cfg.get('times', [])),
                        'timezone':   cfg.get('localTimeZoneId', ''),
                        'weekdays':   ', '.join(cfg.get('weekDays', [])),
                        'interval':   str(cfg.get('interval', '')) if cfg.get('interval') else '',
                        'start_date': cfg.get('startDateTime', '')[:10] if cfg.get('startDateTime') else '',
                        'end_date':   cfg.get('endDateTime', '')[:10]   if cfg.get('endDateTime')   else '',
                    })
                return results
            except Exception:
                pass
    return []


def get_invoke_targets(parts):
    targets = []

    def _extract_name(ref):
        if isinstance(ref, dict):
            return (ref.get('referenceName') or ref.get('pipelineName')
                    or ref.get('name') or '')
        return str(ref) if ref else ''

    def _scan(activities):
        for act in activities:
            atype = act.get('type', '')
            tp    = act.get('typeProperties', {})
            if atype in ('InvokePipeline', 'ExecutePipeline'):
                ref  = tp.get('pipeline', tp.get('referencePipelineName', {}))
                name = _extract_name(ref)
                if name and name not in targets:
                    targets.append(name)
            for key in ('activities', 'ifTrueActivities', 'ifFalseActivities'):
                if key in tp and isinstance(tp[key], list):
                    _scan(tp[key])
            body = tp.get('body', {})
            if isinstance(body, dict) and 'activities' in body:
                _scan(body['activities'])

    for p in parts:
        path = p.get('path', '')
        if path.endswith('.json') and '.platform' not in path and '.schedules' not in path:
            try:
                data       = json.loads(base64.b64decode(p['payload']).decode('utf-8', errors='replace'))
                activities = data.get('properties', {}).get('activities', [])
                _scan(activities)
            except Exception:
                pass

    return targets


def get_last_run(item_id):
    try:
        r = requests.get(
            BASE + '/workspaces/' + WORKSPACE_ID + '/items/' + item_id + '/jobs/instances?$top=1',
            headers=FAB, timeout=TIMEOUT
        )
        if r.status_code == 403:
            return 'No access', 'No access'
        if r.status_code == 404:
            return 'Never run', 'N/A'
        if r.status_code != 200:
            return 'Error', 'Error'
        runs = r.json().get('value', [])
        if not runs:
            return 'Never run', 'N/A'
        last = runs[0]
        return fmt_time(last.get('startTime', '')), last.get('status', 'Unknown')
    except Exception:
        return 'Error', 'Error'


def process_definition_items(item_type, type_label):
    items = requests.get(
        BASE + '/workspaces/' + WORKSPACE_ID + f'/items?type={item_type}',
        headers=FAB, timeout=TIMEOUT
    ).json().get('value', [])

    if not items:
        console.print(f'  [dim]No {type_label} found in this workspace.[/dim]')
        return []

    rows = []
    with make_progress() as progress:
        task = progress.add_task('', total=len(items), item=f'Scanning {type_label}...')
        for item in items:
            name = item['displayName']
            iid  = item['id']
            progress.update(task, item=trunc(name))

            parts     = get_definition_parts(iid)
            schedules = parse_schedule(parts)
            last_run, last_status = get_last_run(iid)
            base_row = {
                'Item Type': type_label, 'Name': name,
                'Last Run': last_run, 'Last Run Status': last_status,
                'Invokes (children)': '',
            }

            if schedules:
                active_count   = sum(1 for s in schedules if s['enabled'])
                disabled_count = len(schedules) - active_count
                tags = []
                if active_count:
                    tags.append(f'[green]{active_count} active schedule{"s" if active_count > 1 else ""}[/green]')
                if disabled_count:
                    tags.append(f'[yellow]{disabled_count} disabled[/yellow]')
                console.print(f'  [dim]{trunc(name, 60)}[/dim]  {"  ·  ".join(tags)}')
                for s in schedules:
                    rows.append({**base_row,
                        'Enabled':        'Yes' if s['enabled'] else 'No (Disabled)',
                        'Schedule Type':  s['type'],    'Run Time(s)':   s['times'],
                        'Timezone':       s['timezone'], 'Weekdays':      s['weekdays'],
                        'Interval':       s['interval'], 'Schedule From': s['start_date'],
                        'Schedule Until': s['end_date'],
                    })
            elif include_unscheduled:
                rows.append({**base_row,
                    'Enabled': 'No schedule', 'Schedule Type': '', 'Run Time(s)': '',
                    'Timezone': '', 'Weekdays': '', 'Interval': '',
                    'Schedule From': '', 'Schedule Until': '',
                })

            progress.advance(task)

    return rows


# ── Initialise result containers ───────────────────────────────────────────────
pipeline_rows = []; parent_child = []
sm_rows       = []
df_rows       = []
nb_rows       = []
sj_rows       = []
active_pl     = []; disabled_pl = []
active_sm     = []
active_df     = []
active_nb     = []
active_sj     = []

# ── Data Pipelines ─────────────────────────────────────────────────────────────
if 'pipelines' in selected_types:
    _t0 = time.time()
    console.rule('[bold]Data Pipelines[/bold]', style='cyan')
    console.print()
    pipelines = requests.get(
        BASE + '/workspaces/' + WORKSPACE_ID + '/items?type=DataPipeline',
        headers=FAB, timeout=TIMEOUT
    ).json().get('value', [])
    console.print(f'  [dim]{len(pipelines)} pipelines found.[/dim]')
    console.print()

    definition_errors = 0
    with make_progress() as progress:
        task = progress.add_task('', total=len(pipelines), item='Initialising...')
        for pl in pipelines:
            name = pl['displayName']
            pid  = pl['id']
            progress.update(task, item=trunc(name))

            if pid in checkpoint:
                c           = checkpoint[pid]
                schedules   = c['schedules']
                invokes     = c['invokes']
                last_run    = c['last_run']
                last_status = c['last_status']
            else:
                parts = get_definition_parts(pid)
                if parts is None:
                    definition_errors += 1
                    parts = []
                schedules   = parse_schedule(parts)
                invokes     = get_invoke_targets(parts)
                last_run, last_status = get_last_run(pid)
                checkpoint[pid] = {
                    'schedules': schedules, 'invokes': invokes,
                    'last_run': last_run, 'last_status': last_status,
                }
                CHECKPOINT_FILE.write_text(json.dumps(checkpoint), encoding='utf-8')

            for child in invokes:
                parent_child.append({'Parent Pipeline': name, 'Child Pipeline': child})

            tags = []
            if schedules:
                active_count   = sum(1 for s in schedules if s['enabled'])
                disabled_count = len(schedules) - active_count
                if active_count:
                    tags.append(f'[green]{active_count} active schedule{"s" if active_count > 1 else ""}[/green]')
                if disabled_count:
                    tags.append(f'[yellow]{disabled_count} disabled[/yellow]')
            if invokes:
                tags.append(f'[cyan]→ {len(invokes)} child{"ren" if len(invokes) > 1 else ""}[/cyan]')
            if tags:
                console.print(f'  [dim]{trunc(name, 60)}[/dim]  {"  ·  ".join(tags)}')

            base_row = {
                'Item Type': 'Data Pipeline', 'Name': name,
                'Last Run': last_run, 'Last Run Status': last_status,
                'Invokes (children)': ', '.join(invokes) if invokes else '',
            }
            if schedules:
                for s in schedules:
                    pipeline_rows.append({**base_row,
                        'Enabled':        'Yes' if s['enabled'] else 'No (Disabled)',
                        'Schedule Type':  s['type'],    'Run Time(s)':   s['times'],
                        'Timezone':       s['timezone'], 'Weekdays':      s['weekdays'],
                        'Interval':       s['interval'], 'Schedule From': s['start_date'],
                        'Schedule Until': s['end_date'],
                    })
            elif include_unscheduled:
                pipeline_rows.append({**base_row,
                    'Enabled': 'No schedule', 'Schedule Type': '', 'Run Time(s)': '',
                    'Timezone': '', 'Weekdays': '', 'Interval': '',
                    'Schedule From': '', 'Schedule Until': '',
                })

            progress.advance(task)

    active_pl   = [r for r in pipeline_rows if r['Enabled'] == 'Yes']
    disabled_pl = [r for r in pipeline_rows if 'Disabled' in r['Enabled']]
    console.print()
    summary_parts = [
        f'[green]{len(active_pl)} active[/green]',
        f'[yellow]{len(disabled_pl)} disabled[/yellow]',
        f'[cyan]{len(parent_child)} parent-child relationships[/cyan]',
    ]
    if definition_errors:
        summary_parts.append(f'[red]{definition_errors} could not be fetched (network error)[/red]')
    console.print(f'  [bold green]✓[/bold green]  ' + '  ·  '.join(summary_parts))
    console.print()
    section_times['pipelines'] = time.time() - _t0

# ── Semantic Models ────────────────────────────────────────────────────────────
if 'semantic_models' in selected_types:
    _t0 = time.time()
    console.rule('[bold]Semantic Models[/bold]', style='cyan')
    console.print()
    try:
        datasets = requests.get(
            PBI_BASE + '/groups/' + WORKSPACE_ID + '/datasets',
            headers=PBI, timeout=TIMEOUT
        ).json().get('value', [])
    except Exception as e:
        console.print(f'  [red]Could not reach the Power BI API: {type(e).__name__}. Check your network connection.[/red]')
        datasets = []
    console.print(f'  [dim]{len(datasets)} semantic models found.[/dim]')
    console.print()

    with make_progress() as progress:
        task = progress.add_task('', total=len(datasets), item='Initialising...')
        for ds in datasets:
            name = ds.get('name', 'Unknown')
            did  = ds['id']
            progress.update(task, item=trunc(name))
            try:
                sched = requests.get(
                    PBI_BASE + '/groups/' + WORKSPACE_ID + '/datasets/' + did + '/refreshSchedule',
                    headers=PBI, timeout=TIMEOUT
                ).json()
            except Exception:
                sched = {}
            try:
                history  = requests.get(
                    PBI_BASE + '/groups/' + WORKSPACE_ID + '/datasets/' + did + '/refreshes?$top=1',
                    headers=PBI, timeout=TIMEOUT
                ).json()
                hist        = history.get('value', [])
                last_run    = fmt_time(hist[0].get('startTime', '')) if hist else 'Never run'
                last_status = hist[0].get('status', 'N/A') if hist else 'N/A'
            except Exception:
                last_run, last_status = 'Error', 'Error'

            enabled = sched.get('enabled', False)
            times   = sched.get('times', [])
            days    = sched.get('days', [])

            if enabled:
                console.print(f'  [dim]{trunc(name, 60)}[/dim]  [green]active refresh[/green]')

            if enabled or include_unscheduled:
                sm_rows.append({
                    'Item Type': 'Semantic Model', 'Name': name,
                    'Enabled':        'Yes' if enabled else 'No schedule',
                    'Schedule Type':  sched.get('frequency', ''),
                    'Run Time(s)':    ', '.join(times),
                    'Timezone':       sched.get('localTimeZoneId', 'UTC'),
                    'Weekdays':       ', '.join(days) if days else '',
                    'Interval': '', 'Schedule From': '', 'Schedule Until': '',
                    'Last Run': last_run, 'Last Run Status': last_status,
                    'Invokes (children)': '',
                })
            progress.advance(task)

    active_sm = [r for r in sm_rows if r['Enabled'] == 'Yes']
    console.print()
    console.print(
        f'  [bold green]✓[/bold green]  '
        f'[green]{len(active_sm)} with active refresh schedule[/green]'
        + (f'  ·  [dim]{len(sm_rows) - len(active_sm)} without[/dim]' if include_unscheduled else '')
    )
    console.print()
    section_times['semantic_models'] = time.time() - _t0

# ── Dataflows ──────────────────────────────────────────────────────────────────
if 'dataflows' in selected_types:
    _t0 = time.time()
    console.rule('[bold]Dataflows[/bold]', style='cyan')
    console.print()
    all_items = requests.get(
        BASE + '/workspaces/' + WORKSPACE_ID + '/items',
        headers=FAB, timeout=TIMEOUT
    ).json().get('value', [])
    dataflows = [it for it in all_items if it.get('type') == 'Dataflow']
    console.print(f'  [dim]{len(dataflows)} dataflows found.[/dim]')

    if not dataflows:
        console.print('  [dim]No dataflows to process.[/dim]')
    else:
        console.print()
        with make_progress() as progress:
            task = progress.add_task('', total=len(dataflows), item='Initialising...')
            for df in dataflows:
                name = df['displayName']
                did  = df['id']
                progress.update(task, item=trunc(name))
                try:
                    sched = requests.get(
                        PBI_BASE + '/groups/' + WORKSPACE_ID + '/dataflows/' + did + '/refreshSchedule',
                        headers=PBI, timeout=TIMEOUT
                    ).json()
                except Exception:
                    sched = {}
                try:
                    history  = requests.get(
                        PBI_BASE + '/groups/' + WORKSPACE_ID + '/dataflows/' + did + '/transactions?$top=1',
                        headers=PBI, timeout=TIMEOUT
                    ).json()
                    hist        = history.get('value', [])
                    last_run    = fmt_time(hist[0].get('startTime', '')) if hist else 'Never run'
                    last_status = hist[0].get('status', 'N/A') if hist else 'N/A'
                except Exception:
                    last_run, last_status = 'Error', 'Error'

                enabled      = sched.get('enabled', False)
                times        = sched.get('times', [])
                days         = sched.get('days', [])
                has_schedule = 'enabled' in sched or 'times' in sched

                if enabled:
                    console.print(f'  [dim]{trunc(name, 60)}[/dim]  [green]active refresh[/green]')

                if enabled or include_unscheduled:
                    df_rows.append({
                        'Item Type': 'Dataflow', 'Name': name,
                        'Enabled':        'Yes' if enabled else ('No schedule' if not has_schedule else 'No (Disabled)'),
                        'Schedule Type':  sched.get('frequency', ''),
                        'Run Time(s)':    ', '.join(times),
                        'Timezone':       sched.get('localTimeZoneId', 'UTC'),
                        'Weekdays':       ', '.join(days) if days else '',
                        'Interval': '', 'Schedule From': '', 'Schedule Until': '',
                        'Last Run': last_run, 'Last Run Status': last_status,
                        'Invokes (children)': '',
                    })
                progress.advance(task)

    active_df = [r for r in df_rows if r['Enabled'] == 'Yes']
    console.print()
    console.print(
        f'  [bold green]✓[/bold green]  '
        f'[green]{len(active_df)} with active refresh schedule[/green]'
        + (f'  ·  [dim]{len(df_rows) - len(active_df)} without[/dim]' if include_unscheduled else '')
    )
    console.print()
    section_times['dataflows'] = time.time() - _t0

# ── Notebooks ──────────────────────────────────────────────────────────────────
if 'notebooks' in selected_types:
    _t0 = time.time()
    console.rule('[bold]Notebooks[/bold]', style='cyan')
    console.print()
    nb_rows   = process_definition_items('Notebook', 'Notebook')
    active_nb = [r for r in nb_rows if r['Enabled'] == 'Yes']
    console.print()
    console.print(
        f'  [bold green]✓[/bold green]  '
        f'[green]{len(active_nb)} with active schedule[/green]'
    )
    console.print()
    section_times['notebooks'] = time.time() - _t0

# ── Spark Job Definitions ──────────────────────────────────────────────────────
if 'spark_jobs' in selected_types:
    _t0 = time.time()
    console.rule('[bold]Spark Job Definitions[/bold]', style='cyan')
    console.print()
    sj_rows   = process_definition_items('SparkJobDefinition', 'Spark Job Definition')
    active_sj = [r for r in sj_rows if r['Enabled'] == 'Yes']
    console.print()
    console.print(
        f'  [bold green]✓[/bold green]  '
        f'[green]{len(active_sj)} with active schedule[/green]'
    )
    console.print()
    section_times['spark_jobs'] = time.time() - _t0

# ── Build Excel ────────────────────────────────────────────────────────────────
console.rule('[bold]Building Excel[/bold]', style='cyan')
console.print()

NAVY     = 'FF1F3864'
BLUE     = 'FF2E75B6'
WHITE    = 'FFFFFFFF'
LIGHT    = 'FFD6E4F7'
GREY     = 'FFF2F2F2'
GREEN_BG = 'FFE2EFDA'
GREEN_FG = 'FF375623'
RED_BG   = 'FFFFC7CE'
RED_FG   = 'FF9C0006'
AMBER_BG = 'FFFFFF99'
AMBER_FG = 'FF833C00'
thin = Side(style='thin', color='FFD9D9D9')
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = [
    'Item Type', 'Name', 'Enabled', 'Schedule Type', 'Run Time(s)', 'Timezone',
    'Weekdays', 'Interval', 'Schedule From', 'Schedule Until',
    'Last Run', 'Last Run Status', 'Invokes (children)',
]
PARENT_COLS = ['Parent Pipeline', 'Child Pipeline']

stamp      = f'{WORKSPACE_NAME}  |  Extracted: ' + datetime.now().strftime('%Y-%m-%d %H:%M UTC')
all_active = active_pl + active_sm + active_df + active_nb + active_sj


def make_header(ws, title, subtitle, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    c = ws.cell(1, 1)
    c.value = title
    c.font  = Font(name='Arial', bold=True, color=WHITE, size=13)
    c.fill  = PatternFill('solid', start_color=NAVY)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
    c = ws.cell(2, 1)
    c.value = subtitle
    c.font  = Font(name='Arial', color=WHITE, size=10)
    c.fill  = PatternFill('solid', start_color=BLUE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18

    for col, h in enumerate(cols, 1):
        c = ws.cell(3, col)
        c.value = h
        c.font  = Font(name='Arial', bold=True, color=WHITE, size=10)
        c.fill  = PatternFill('solid', start_color=NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = bdr
    ws.row_dimensions[3].height = 28


def write_rows(ws, rows, cols):
    today = datetime.now().strftime('%Y-%m-%d')
    for i, row in enumerate(rows):
        r   = 4 + i
        alt = (i % 2 == 0)
        for col, key in enumerate(cols, 1):
            c   = ws.cell(r, col)
            val = row.get(key, '')
            c.value     = val
            c.font      = Font(name='Arial', size=9)
            c.alignment = Alignment(vertical='center', wrap_text=(col <= 2))
            c.border    = bdr

            if key == 'Enabled':
                if val == 'Yes':
                    c.font = Font(name='Arial', size=9, bold=True, color=GREEN_FG)
                    c.fill = PatternFill('solid', start_color=GREEN_BG)
                elif 'Disabled' in str(val):
                    c.font = Font(name='Arial', size=9, color='FF7F7F7F')
                    c.fill = PatternFill('solid', start_color=GREY)
                else:
                    c.fill = PatternFill('solid', start_color=LIGHT if alt else WHITE)
            elif key == 'Last Run Status':
                if val in ('Succeeded', 'Completed'):
                    c.font = Font(name='Arial', size=9, color=GREEN_FG)
                    c.fill = PatternFill('solid', start_color=GREEN_BG)
                elif val == 'Failed':
                    c.font = Font(name='Arial', size=9, bold=True, color=RED_FG)
                    c.fill = PatternFill('solid', start_color=RED_BG)
                elif val == 'InProgress':
                    c.font = Font(name='Arial', size=9, color=AMBER_FG)
                    c.fill = PatternFill('solid', start_color=AMBER_BG)
                else:
                    c.fill = PatternFill('solid', start_color=LIGHT if alt else WHITE)
            elif key == 'Schedule Until':
                if val and val < today:
                    c.font = Font(name='Arial', size=9, bold=True, color=AMBER_FG)
                    c.fill = PatternFill('solid', start_color=AMBER_BG)
                else:
                    c.fill = PatternFill('solid', start_color=LIGHT if alt else WHITE)
            else:
                c.fill = PatternFill('solid', start_color=LIGHT if alt else WHITE)

        ws.row_dimensions[r].height = 18


def set_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


wb     = openpyxl.Workbook()
widths = [14, 40, 14, 13, 14, 24, 16, 10, 13, 13, 22, 16, 40]

ws_active = wb.active
ws_active.title = 'Active Schedules'
make_header(ws_active, f'{WORKSPACE_NAME}  |  All Active Schedules',
    stamp + f'  |  {len(all_active)} active scheduled items', COLS)
write_rows(ws_active, all_active, COLS)
set_widths(ws_active, widths)

if 'pipelines' in selected_types:
    ws = wb.create_sheet('Pipelines')
    make_header(ws, f'{WORKSPACE_NAME}  |  Data Pipelines',
        stamp + f'  |  {len(pipeline_rows)} shown  |  {len(active_pl)} active  |  {len(disabled_pl)} disabled', COLS)
    write_rows(ws, pipeline_rows, COLS)
    set_widths(ws, widths)

if 'semantic_models' in selected_types:
    ws = wb.create_sheet('Semantic Models')
    make_header(ws, f'{WORKSPACE_NAME}  |  Semantic Models',
        stamp + f'  |  {len(sm_rows)} shown  |  {len(active_sm)} with active refresh', COLS)
    write_rows(ws, sm_rows, COLS)
    set_widths(ws, widths)

if 'dataflows' in selected_types:
    ws = wb.create_sheet('Dataflows')
    make_header(ws, f'{WORKSPACE_NAME}  |  Dataflows',
        stamp + f'  |  {len(df_rows)} shown  |  {len(active_df)} with active refresh', COLS)
    write_rows(ws, df_rows, COLS)
    set_widths(ws, widths)

if 'notebooks' in selected_types:
    ws = wb.create_sheet('Notebooks')
    make_header(ws, f'{WORKSPACE_NAME}  |  Notebooks',
        stamp + f'  |  {len(nb_rows)} shown  |  {len(active_nb)} with active schedule', COLS)
    write_rows(ws, nb_rows, COLS)
    set_widths(ws, widths)

if 'spark_jobs' in selected_types:
    ws = wb.create_sheet('Spark Job Definitions')
    make_header(ws, f'{WORKSPACE_NAME}  |  Spark Job Definitions',
        stamp + f'  |  {len(sj_rows)} shown  |  {len(active_sj)} with active schedule', COLS)
    write_rows(ws, sj_rows, COLS)
    set_widths(ws, widths)

if 'pipelines' in selected_types:
    ws = wb.create_sheet('Pipeline Relationships')
    make_header(ws, f'{WORKSPACE_NAME}  |  Pipeline Parent-Child Relationships',
        stamp + f'  |  {len(parent_child)} invoke relationships mapped', PARENT_COLS)
    write_rows(ws, parent_child, PARENT_COLS)
    set_widths(ws, [50, 50])

OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(OUTPUT_PATH))

if CHECKPOINT_FILE.exists():
    CHECKPOINT_FILE.unlink()

# ── Summary ────────────────────────────────────────────────────────────────────
console.print(f'  [bold green]✓[/bold green] Excel file saved.')
console.print()


def fmt_dur(seconds):
    seconds = int(seconds)
    h, rem  = divmod(seconds, 3600)
    m, s    = divmod(rem, 60)
    if h:
        return f'{h}h {m}m {s}s'
    if m:
        return f'{m}m {s}s'
    return f'{s}s'


total_elapsed = time.time() - run_start

summary = Table(box=box.ROUNDED, border_style='cyan', show_header=True, header_style='bold white on #1F3864')
summary.add_column('Item Type',  style='dim',        min_width=26)
summary.add_column('Active',     style='bold green', justify='right', min_width=8)
summary.add_column('Inactive',   style='yellow',     justify='right', min_width=10)
summary.add_column('In Output',  style='dim',        justify='right', min_width=10)
summary.add_column('Time taken', style='cyan',       justify='right', min_width=12)

if 'pipelines' in selected_types:
    summary.add_row('Data Pipelines', str(len(active_pl)), str(len(disabled_pl)),
                    str(len(pipeline_rows)), fmt_dur(section_times.get('pipelines', 0)))
if 'semantic_models' in selected_types:
    summary.add_row('Semantic Models', str(len(active_sm)), str(len(sm_rows) - len(active_sm)),
                    str(len(sm_rows)), fmt_dur(section_times.get('semantic_models', 0)))
if 'dataflows' in selected_types:
    summary.add_row('Dataflows', str(len(active_df)), str(len(df_rows) - len(active_df)),
                    str(len(df_rows)), fmt_dur(section_times.get('dataflows', 0)))
if 'notebooks' in selected_types:
    summary.add_row('Notebooks', str(len(active_nb)), str(len(nb_rows) - len(active_nb)),
                    str(len(nb_rows)), fmt_dur(section_times.get('notebooks', 0)))
if 'spark_jobs' in selected_types:
    summary.add_row('Spark Job Definitions', str(len(active_sj)), str(len(sj_rows) - len(active_sj)),
                    str(len(sj_rows)), fmt_dur(section_times.get('spark_jobs', 0)))
summary.add_section()
summary.add_row('[bold]Total active schedules[/bold]', f'[bold green]{len(all_active)}[/bold green]', '', '', '')
if 'pipelines' in selected_types:
    summary.add_row('[bold]Pipeline relationships[/bold]', f'[bold cyan]{len(parent_child)}[/bold cyan]', '', '', '')
summary.add_section()
summary.add_row('[bold]Total run time[/bold]', '', '', '', f'[bold cyan]{fmt_dur(total_elapsed)}[/bold cyan]')

console.print(summary)

if all_active:
    dates_table = Table(
        box=box.ROUNDED, border_style='cyan', show_header=True,
        header_style='bold white on #1F3864',
        title='[bold]Active Schedule Windows[/bold]',
    )
    dates_table.add_column('Name',       style='bold',   min_width=30)
    dates_table.add_column('Type',       style='dim',    min_width=20)
    dates_table.add_column('Start Date', style='green',  min_width=13, justify='center')
    dates_table.add_column('End Date',   style='yellow', min_width=13, justify='center')
    for row in sorted(all_active, key=lambda r: (r['Item Type'], r['Name'])):
        dates_table.add_row(
            row['Name'],
            row['Item Type'],
            row.get('Schedule From') or '[dim]not set[/dim]',
            row.get('Schedule Until') or '[dim]not set[/dim]',
        )
    console.print()
    console.print(dates_table)

console.print()
console.print(f'  [bold]Saved to:[/bold] [cyan]{OUTPUT_PATH}[/cyan]')
console.print()
