"""
Fabric Schedule Disabler
=========================
Connects to a Fabric workspace, finds all items with active schedules, and
disables them. Works in two modes:

  scan mode (default): live API scan, finds active schedules right now
  file mode:           reads the Excel produced by schedule_extractor.py

Stage guardrails (auto-detected from workspace name):
  DEV   - proceeds after a single confirmation prompt
  UAT   - requires typing the workspace name to confirm, no auto/watch mode
  PROD  - refuses to run entirely

Supports:
  Data Pipelines, Notebooks, Spark Job Definitions  (Fabric definition API)
  Semantic Models                                    (Power BI refresh schedule API)
  Dataflows                                         (Power BI dataflow refresh API)

Nothing is changed until you confirm. All changes are logged in the result table.
"""

# Author: Abigail Woolley, AmaliTech
# Built with AI assistance

import sys
import subprocess
from pathlib import Path

REQUIRED = ['msal', 'requests', 'openpyxl', 'rich']
for pkg in REQUIRED:
    try:
        __import__(pkg)
    except ImportError:  # pragma: no cover
        print(f'Installing {pkg}...')
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg, '--quiet'])
        print(f'  {pkg} installed.')

import io
import msal
import requests
import json
import base64
import time
from datetime import datetime
from openpyxl import load_workbook
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich import box

try:
    sys.stdout.reconfigure(encoding='utf-8')
except (AttributeError, io.UnsupportedOperation):  # pragma: no cover
    pass
console = Console()

# ── Static configuration ───────────────────────────────────────────────────────
AUTHORITY  = 'https://login.microsoftonline.com/organizations'
TIMEOUT    = 60
CACHE_FILE = Path.home() / '.fabric_token_cache.bin'
FAB_BASE   = 'https://api.fabric.microsoft.com/v1'
PBI_BASE   = 'https://api.powerbi.com/v1.0/myorg'
FAB_SCOPES = ['https://api.fabric.microsoft.com/.default']
PBI_SCOPES = ['https://analysis.windows.net/powerbi/api/.default']

DEFINITION_TYPES = {'Data Pipeline', 'Notebook', 'Spark Job Definition'}
FABRIC_TYPE_MAP  = {
    'Data Pipeline':        'DataPipeline',
    'Notebook':             'Notebook',
    'Spark Job Definition': 'SparkJobDefinition',
}

# ── Runtime globals (set by main, read by helper functions) ───────────────────
WORKSPACE_ID: str = ''
FAB: dict         = {}
PBI: dict         = {}


# ── Stage detection ────────────────────────────────────────────────────────────
def detect_stage(workspace_name):
    ws_lower = workspace_name.lower()
    if any(p in ws_lower for p in ('uat', 'staging', 'stg', 'preprod', 'pre-prod')):
        return 'UAT'
    if any(p in ws_lower for p in ('prod', 'production', 'prd')):
        return 'PROD'
    return 'DEV'


# ── Auth helper ────────────────────────────────────────────────────────────────
def get_token(scopes, app, cache):
    accounts = app.get_accounts()
    token    = None
    if accounts:
        token = app.acquire_token_silent(scopes, account=accounts[0])
    if not token:
        console.print('[yellow]Opening browser for Microsoft sign-in...[/yellow]')
        try:
            token = app.acquire_token_interactive(scopes)
        except Exception:
            flow = app.initiate_device_flow(scopes)
            if 'error' in flow:
                console.print(f'[bold red]Device flow error: {flow.get("error_description", flow["error"])}[/bold red]')
                sys.exit(1)
            console.print(flow['message'])
            token = app.acquire_token_by_device_flow(flow)
    if not token or 'access_token' not in token:
        console.print('[bold red]ERROR: Authentication failed.[/bold red]')
        console.print((token or {}).get('error_description', 'Unknown error'))
        sys.exit(1)
    if cache.has_state_changed:
        CACHE_FILE.write_text(cache.serialize(), encoding='utf-8')
        try:
            CACHE_FILE.chmod(0o600)
        except (AttributeError, NotImplementedError):
            pass
    return token['access_token']


# ── Fabric definition API helpers ─────────────────────────────────────────────
def get_parts(item_id):
    r = requests.post(
        f'{FAB_BASE}/workspaces/{WORKSPACE_ID}/items/{item_id}/getDefinition',
        headers=FAB, timeout=TIMEOUT,
    )
    if r.status_code == 202:
        op_url = r.headers.get('Location') or r.headers.get('location')
        if not op_url:
            return []
        for _ in range(15):
            time.sleep(2)
            poll = requests.get(op_url, headers=FAB, timeout=TIMEOUT)
            if poll.status_code == 200:
                return poll.json().get('definition', {}).get('parts', [])
        return []
    return r.json().get('definition', {}).get('parts', []) if r.ok else []


def push_parts(item_id, parts):
    r = requests.post(
        f'{FAB_BASE}/workspaces/{WORKSPACE_ID}/items/{item_id}/updateDefinition',
        headers=FAB, timeout=TIMEOUT,
        json={'definition': {'parts': parts}},
    )
    if r.status_code == 202:
        op_url = r.headers.get('Location') or r.headers.get('location')
        if not op_url:
            return False
        for _ in range(15):
            time.sleep(2)
            poll = requests.get(op_url, headers=FAB, timeout=TIMEOUT)
            s = poll.json().get('status', '')
            if s in ('Succeeded', 'Failed') or poll.status_code == 200:
                return s != 'Failed'
        return False
    return r.status_code in (200, 202, 204)


def disable_fabric_schedule(item_id):
    parts = get_parts(item_id)
    if not parts:
        return False, 'Could not fetch definition'
    updated   = False
    new_parts = []
    for part in parts:
        if part.get('path') == '.schedules':
            try:
                data = json.loads(base64.b64decode(part['payload']).decode('utf-8'))
                for s in data.get('schedules', []):
                    if s.get('enabled'):
                        s['enabled'] = False
                        updated = True
                new_payload = base64.b64encode(
                    json.dumps(data, indent=2).encode('utf-8')
                ).decode('utf-8')
                new_parts.append({**part, 'payload': new_payload})
            except Exception as e:
                return False, f'Parse error: {e}'
        else:
            new_parts.append(part)
    if not updated:
        return True, 'Already disabled'
    ok = push_parts(item_id, new_parts)
    return ok, 'Disabled' if ok else 'API error on update'


# ── Power BI schedule helper ───────────────────────────────────────────────────
def disable_pbi_schedule(endpoint):
    r = requests.get(endpoint, headers=PBI, timeout=TIMEOUT)
    if not r.ok:
        return False, f'Could not fetch schedule ({r.status_code})'
    sched = r.json()
    if not sched.get('enabled', False):
        return True, 'Already disabled'
    sched['enabled'] = False
    patch = requests.patch(endpoint, headers=PBI, timeout=TIMEOUT, json={'value': sched})
    return patch.ok, 'Disabled' if patch.ok else f'API error {patch.status_code}'


# ── Entry point ────────────────────────────────────────────────────────────────
def main():  # pragma: no cover
    global WORKSPACE_ID, FAB, PBI

    # Load CLIENT_ID
    sys.path.insert(0, str(Path(__file__).parent.parent))
    try:
        from config import CLIENT_ID  # noqa: PLC0415
    except ImportError:
        console.print('[bold red]ERROR: config.py not found.[/bold red]')
        console.print('Copy config.example.py to config.py and add your CLIENT_ID.')
        sys.exit(1)

    # Banner
    console.print()
    console.print(Panel.fit(
        '[bold cyan]Fabric Schedule Disabler[/bold cyan]',
        border_style='cyan',
    ))
    console.print()

    # Mode selection
    console.print('[bold]How would you like to find active schedules?[/bold]')
    console.print()
    console.print('  [[bold cyan]1[/bold cyan]] Live scan  [dim](connect to workspace now and find active schedules)[/dim]')
    console.print('  [[bold cyan]2[/bold cyan]] From file  [dim](use an Excel file from schedule_extractor.py)[/dim]')
    console.print()
    mode_raw = console.input('Enter [bold]1[/bold] or [bold]2[/bold] [dim](default: 1)[/dim]: ').strip()
    use_file = mode_raw == '2'
    console.print()

    # Auth
    cache = msal.SerializableTokenCache()
    if CACHE_FILE.exists():
        cache.deserialize(CACHE_FILE.read_text(encoding='utf-8'))
    app = msal.PublicClientApplication(CLIENT_ID, authority=AUTHORITY, token_cache=cache)

    console.print('[dim]Authenticating...[/dim]')
    fab_at = get_token(FAB_SCOPES, app, cache)
    pbi_at = get_token(PBI_SCOPES, app, cache)
    FAB = {'Authorization': 'Bearer ' + fab_at, 'Content-Type': 'application/json'}
    PBI = {'Authorization': 'Bearer ' + pbi_at, 'Content-Type': 'application/json'}
    console.print('[bold green]✓[/bold green] Authenticated.')
    console.print()

    # Workspace selection
    console.print('[dim]Fetching available workspaces...[/dim]')
    ws_resp = requests.get(FAB_BASE + '/workspaces', headers=FAB, timeout=TIMEOUT)
    if ws_resp.status_code != 200:
        console.print(f'[bold red]ERROR: Could not list workspaces ({ws_resp.status_code})[/bold red]')
        sys.exit(1)
    all_workspaces = ws_resp.json().get('value', [])
    console.print(f'[bold green]✓[/bold green] {len(all_workspaces)} workspaces available.')
    console.print()

    search      = console.input('[bold]Enter workspace name[/bold] [dim](or part of it)[/dim]: ').strip().lower()
    search_norm = search.replace(' ', '').replace('-', '')
    matches     = [
        w for w in all_workspaces
        if search in w['displayName'].lower()
        or search_norm in w['displayName'].lower().replace(' ', '').replace('-', '')
    ]

    if not matches:
        console.print(f'[red]No workspace found matching "[bold]{search}[/bold]".[/red]')
        sys.exit(1)

    if len(matches) == 1:
        selected_workspace = matches[0]
    else:
        console.print('\n[yellow]Multiple matches found:[/yellow]')
        for i, w in enumerate(matches, 1):
            console.print(f'  [[bold]{i}[/bold]] {w["displayName"]}')
        choice = console.input('Enter number: ').strip()
        try:
            selected_workspace = matches[int(choice) - 1]
        except (ValueError, IndexError):
            console.print('[red]Invalid choice.[/red]')
            sys.exit(1)

    WORKSPACE_ID   = selected_workspace['id']
    workspace_name = selected_workspace['displayName']
    console.print()
    console.print(f'  [bold]Workspace:[/bold] [cyan]{workspace_name}[/cyan]')
    console.print()

    # Stage guardrails
    stage = detect_stage(workspace_name)

    if stage == 'PROD':
        console.print('[bold red]PROD workspace detected.[/bold red]')
        console.print('[red]This tool will not disable schedules in a production workspace.[/red]')
        console.print('[dim]If you need to change production schedules, do it directly in the Fabric portal.[/dim]')
        sys.exit(0)

    if stage == 'UAT':
        console.print('[bold yellow]UAT workspace detected.[/bold yellow]')
        console.print('[yellow]Extra confirmation required before any changes are made.[/yellow]')
        console.print()
    else:
        console.print('[bold green]DEV workspace detected.[/bold green]')
        console.print()

    # Build list of items to disable
    to_disable = {}

    if use_file:
        downloads  = Path.home() / 'Downloads'
        xlsx_files = sorted(downloads.glob('* - Schedule Inventory.xlsx'), key=lambda f: f.stat().st_mtime, reverse=True)

        if not xlsx_files:
            console.print('[red]No Schedule Inventory Excel found in Downloads.[/red]')
            console.print('[dim]Run schedule_extractor.py first to generate one.[/dim]')
            sys.exit(1)

        if len(xlsx_files) == 1:
            excel_path = xlsx_files[0]
        else:
            console.print('[yellow]Multiple inventory files found:[/yellow]')
            for i, f in enumerate(xlsx_files[:5], 1):
                age = int((datetime.now().timestamp() - f.stat().st_mtime) / 60)
                console.print(f'  [[bold]{i}[/bold]] {f.name}  [dim]({age} min ago)[/dim]')
            choice = console.input('Enter number [dim](default: 1)[/dim]: ').strip()
            try:
                excel_path = xlsx_files[int(choice) - 1] if choice else xlsx_files[0]
            except (ValueError, IndexError):
                excel_path = xlsx_files[0]

        console.print(f'[dim]Reading:[/dim] {excel_path.name}')
        wb = load_workbook(excel_path, read_only=True, data_only=True)

        target_sheet = None
        header_row   = None
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
                if row and 'Enabled' in row and 'Name' in row and 'Item Type' in row:
                    target_sheet = ws
                    header_row   = row_idx
                    break
            if target_sheet:
                break

        if not target_sheet:
            console.print('[red]Could not find a sheet with Enabled/Name/Item Type columns.[/red]')
            sys.exit(1)

        headers = list(next(target_sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
        col     = {h: i for i, h in enumerate(headers) if h}

        for row in target_sheet.iter_rows(min_row=header_row + 1, values_only=True):
            enabled   = str(row[col['Enabled']] or '').strip()
            item_type = str(row[col['Item Type']] or '').strip()
            name      = str(row[col['Name']] or '').strip()
            if enabled == 'Yes' and item_type and name:
                to_disable[(item_type, name)] = True
        wb.close()
        console.print(f'[bold green]✓[/bold green] {len(to_disable)} active schedule(s) found in file.')

    else:
        # Live scan
        console.print('[dim]Scanning workspace for active schedules...[/dim]')
        console.print()
        scan_errors = 0

        pipelines = requests.get(
            f'{FAB_BASE}/workspaces/{WORKSPACE_ID}/items?type=DataPipeline',
            headers=FAB, timeout=TIMEOUT
        ).json().get('value', [])

        console.print(f'  [dim]Checking {len(pipelines)} pipelines...[/dim]')
        for pl in pipelines:
            name = pl['displayName']
            pid  = pl['id']
            try:
                r = requests.get(
                    f'{FAB_BASE}/workspaces/{WORKSPACE_ID}/items/{pid}/jobs/schedules',
                    headers=FAB, timeout=TIMEOUT
                )
                if r.ok and any(s.get('enabled') for s in r.json().get('value', [])):
                    to_disable[('Data Pipeline', name)] = True
                    console.print(f'    [green]+[/green] {name}')
            except Exception:
                scan_errors += 1

        for item_type_label, api_type in [('Notebook', 'Notebook'), ('Spark Job Definition', 'SparkJobDefinition')]:
            fab_items = requests.get(
                f'{FAB_BASE}/workspaces/{WORKSPACE_ID}/items?type={api_type}',
                headers=FAB, timeout=TIMEOUT
            ).json().get('value', [])
            console.print(f'  [dim]Checking {len(fab_items)} {item_type_label.lower()}s...[/dim]')
            for it in fab_items:
                name = it['displayName']
                iid  = it['id']
                try:
                    r = requests.get(
                        f'{FAB_BASE}/workspaces/{WORKSPACE_ID}/items/{iid}/jobs/schedules',
                        headers=FAB, timeout=TIMEOUT
                    )
                    if r.ok and any(s.get('enabled') for s in r.json().get('value', [])):
                        to_disable[(item_type_label, name)] = True
                        console.print(f'    [green]+[/green] {name}')
                except Exception:
                    scan_errors += 1

        datasets = requests.get(
            f'{PBI_BASE}/groups/{WORKSPACE_ID}/datasets',
            headers=PBI, timeout=TIMEOUT
        ).json().get('value', [])

        console.print(f'  [dim]Checking {len(datasets)} semantic models...[/dim]')
        for ds in datasets:
            name = ds.get('name', '')
            did  = ds['id']
            try:
                sched = requests.get(
                    f'{PBI_BASE}/groups/{WORKSPACE_ID}/datasets/{did}/refreshSchedule',
                    headers=PBI, timeout=TIMEOUT
                ).json()
                if sched.get('enabled', False):
                    to_disable[('Semantic Model', name)] = True
                    console.print(f'    [green]+[/green] {name}')
            except Exception:
                scan_errors += 1

        all_items = requests.get(
            f'{FAB_BASE}/workspaces/{WORKSPACE_ID}/items',
            headers=FAB, timeout=TIMEOUT
        ).json().get('value', [])
        dataflows = [it for it in all_items if it.get('type') == 'Dataflow']

        if dataflows:
            console.print(f'  [dim]Checking {len(dataflows)} dataflows...[/dim]')
            for df in dataflows:
                name = df['displayName']
                did  = df['id']
                try:
                    sched = requests.get(
                        f'{PBI_BASE}/groups/{WORKSPACE_ID}/dataflows/{did}/refreshSchedule',
                        headers=PBI, timeout=TIMEOUT
                    ).json()
                    if sched.get('enabled', False):
                        to_disable[('Dataflow', name)] = True
                        console.print(f'    [green]+[/green] {name}')
                except Exception:
                    scan_errors += 1

        console.print()
        console.print(f'[bold green]✓[/bold green] Scan complete. {len(to_disable)} active schedule(s) found.')
        if scan_errors:
            console.print(f'  [yellow]⚠[/yellow] {scan_errors} item(s) could not be checked (network or permission error).')

    console.print()

    if not to_disable:
        console.print('[yellow]No active schedules found. Nothing to disable.[/yellow]')
        sys.exit(0)

    # Preview table
    preview = Table(
        title=f'[bold]{len(to_disable)} active schedule(s) to disable[/bold]',
        box=box.ROUNDED, border_style='yellow', show_header=True,
        header_style='bold white on #1F3864',
    )
    preview.add_column('Item Type', style='cyan', min_width=20)
    preview.add_column('Name',      style='white', min_width=40)
    for (itype, name) in sorted(to_disable.keys()):
        preview.add_row(itype, name)
    console.print(preview)
    console.print()

    # Confirmation
    if stage == 'UAT':
        console.print('[bold yellow]UAT confirmation required.[/bold yellow]')
        console.print(f'Type the workspace name exactly to proceed: [bold]{workspace_name}[/bold]')
        typed = console.input('Workspace name: ').strip()
        if typed != workspace_name:
            console.print('[red]Name did not match. Cancelled.[/red]')
            sys.exit(0)
    else:
        confirm = console.input('[bold yellow]Proceed and disable all of the above? [y/N]: [/bold yellow]').strip().lower()
        if confirm not in ('y', 'yes'):
            console.print('[dim]Cancelled.[/dim]')
            sys.exit(0)

    console.print()

    # Build ID lookup maps
    fab_id_map = {}
    for fab_type in FABRIC_TYPE_MAP.values():
        r = requests.get(
            f'{FAB_BASE}/workspaces/{WORKSPACE_ID}/items?type={fab_type}',
            headers=FAB, timeout=TIMEOUT,
        )
        if r.ok:
            for item in r.json().get('value', []):
                fab_id_map[(fab_type, item['displayName'].lower())] = item['id']

    pbi_dataset_map = {}
    r = requests.get(f'{PBI_BASE}/groups/{WORKSPACE_ID}/datasets', headers=PBI, timeout=TIMEOUT)
    if r.ok:
        for ds in r.json().get('value', []):
            pbi_dataset_map[ds['name'].lower()] = ds['id']

    pbi_dataflow_map = {}
    r = requests.get(f'{PBI_BASE}/groups/{WORKSPACE_ID}/dataflows', headers=PBI, timeout=TIMEOUT)
    if r.ok:
        for df in r.json().get('value', []):
            pbi_dataflow_map[df.get('name', df.get('modelName', '')).lower()] = df['objectId']

    # Disable loop
    results = []

    for (item_type, name) in sorted(to_disable.keys()):
        name_l = name.lower()

        if item_type in DEFINITION_TYPES:
            fab_type = FABRIC_TYPE_MAP[item_type]
            item_id  = fab_id_map.get((fab_type, name_l))
            if not item_id:
                results.append((item_type, name, False, 'Not found in workspace'))
                continue
            ok, msg = disable_fabric_schedule(item_id)

        elif item_type == 'Semantic Model':
            item_id = pbi_dataset_map.get(name_l)
            if not item_id:
                results.append((item_type, name, False, 'Not found in workspace'))
                continue
            endpoint = f'{PBI_BASE}/groups/{WORKSPACE_ID}/datasets/{item_id}/refreshSchedule'
            ok, msg  = disable_pbi_schedule(endpoint)

        elif item_type == 'Dataflow':
            item_id = pbi_dataflow_map.get(name_l)
            if not item_id:
                results.append((item_type, name, False, 'Not found in workspace'))
                continue
            endpoint = f'{PBI_BASE}/groups/{WORKSPACE_ID}/dataflows/{item_id}/refreshSchedule'
            ok, msg  = disable_pbi_schedule(endpoint)

        else:
            results.append((item_type, name, False, 'Item type not supported'))
            continue

        results.append((item_type, name, ok, msg))
        icon = '[bold green]✓[/bold green]' if ok else '[bold red]✗[/bold red]'
        console.print(f'  {icon}  [cyan]{item_type}[/cyan]  {name}  [dim]{msg}[/dim]')

    # Result summary
    console.print()
    success = sum(1 for _, _, ok, _ in results if ok)
    failed  = len(results) - success

    result_table = Table(
        title='Result Summary',
        box=box.ROUNDED, border_style='cyan', show_header=True,
        header_style='bold white on #1F3864',
    )
    result_table.add_column('Item Type', style='cyan',  min_width=20)
    result_table.add_column('Name',      style='white', min_width=40)
    result_table.add_column('Result',    min_width=20)
    for itype, name, ok, msg in results:
        status = f'[bold green]{msg}[/bold green]' if ok else f'[bold red]{msg}[/bold red]'
        result_table.add_row(itype, name, status)
    console.print(result_table)
    console.print()
    console.print(
        f'  [bold green]{success} disabled[/bold green]'
        + (f'   [bold red]{failed} failed[/bold red]' if failed else '')
    )
    console.print()


if __name__ == '__main__':
    main()
