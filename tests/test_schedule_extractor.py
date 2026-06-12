"""Tests for tools/schedule_extractor.py

Covers:
  - Pure utility functions (trunc, fmt_time, fmt_dur, sanitize_filename)
  - Schedule parsing (parse_schedule)
  - Pipeline child detection (get_invoke_targets)
  - API helpers with mocked HTTP (get_definition_parts, get_last_run)
  - Auth helper (get_token) with mocked MSAL
"""

import base64
import json
import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

# Make tools/ importable without a config.py present
sys.path.insert(0, str(Path(__file__).parent.parent))

# Import the module without triggering main()
import tools.schedule_extractor as ext


# ── trunc ─────────────────────────────────────────────────────────────────────

class TestTrunc:
    def test_short_string_unchanged(self):
        assert ext.trunc('hello') == 'hello'

    def test_exact_limit_unchanged(self):
        s = 'x' * 52
        assert ext.trunc(s) == s

    def test_long_string_truncated(self):
        s = 'a' * 60
        result = ext.trunc(s)
        assert result.endswith('...')
        assert len(result) == 52

    def test_custom_limit(self):
        result = ext.trunc('hello world', n=8)
        assert result == 'hello...'
        assert len(result) == 8

    def test_empty_string(self):
        assert ext.trunc('') == ''


# ── sanitize_filename ─────────────────────────────────────────────────────────

class TestSanitizeFilename:
    def test_clean_name_unchanged(self):
        assert ext.sanitize_filename('GDW DEV') == 'GDW DEV'

    def test_forward_slash_replaced(self):
        assert ext.sanitize_filename('GDW/DEV') == 'GDW_DEV'

    def test_backslash_replaced(self):
        assert ext.sanitize_filename('GDW\\DEV') == 'GDW_DEV'

    def test_colon_replaced(self):
        assert ext.sanitize_filename('GDW:DEV') == 'GDW_DEV'

    def test_all_invalid_chars_replaced(self):
        result = ext.sanitize_filename('a/b\\c:d*e?f"g<h>i|j')
        for ch in r'/\:*?"<>|':
            assert ch not in result
        assert '_' in result

    def test_leading_trailing_spaces_stripped(self):
        assert ext.sanitize_filename('  GDW DEV  ') == 'GDW DEV'


# ── fmt_time ──────────────────────────────────────────────────────────────────

class TestFmtTime:
    def test_iso_string_with_z(self):
        result = ext.fmt_time('2024-03-15T08:30:00Z')
        assert result == '2024-03-15 08:30 UTC'

    def test_iso_string_with_offset(self):
        result = ext.fmt_time('2024-03-15T08:30:00+00:00')
        assert result == '2024-03-15 08:30 UTC'

    def test_empty_string_returns_never(self):
        assert ext.fmt_time('') == 'Never'

    def test_none_returns_never(self):
        assert ext.fmt_time(None) == 'Never'

    def test_never_passthrough(self):
        assert ext.fmt_time('Never') == 'Never'

    def test_invalid_string_returned_as_is(self):
        assert ext.fmt_time('not-a-date') == 'not-a-date'


# ── fmt_dur ───────────────────────────────────────────────────────────────────

class TestFmtDur:
    def test_seconds_only(self):
        assert ext.fmt_dur(45) == '45s'

    def test_minutes_and_seconds(self):
        assert ext.fmt_dur(125) == '2m 5s'

    def test_hours_minutes_seconds(self):
        assert ext.fmt_dur(3661) == '1h 1m 1s'

    def test_zero(self):
        assert ext.fmt_dur(0) == '0s'

    def test_exactly_one_hour(self):
        assert ext.fmt_dur(3600) == '1h 0m 0s'

    def test_float_truncated(self):
        assert ext.fmt_dur(90.9) == '1m 30s'


# ── parse_schedule ────────────────────────────────────────────────────────────

class TestParseSchedule:
    def test_active_schedule_parsed(self, schedule_part_active):
        result = ext.parse_schedule(schedule_part_active)
        assert len(result) == 1
        assert result[0]['enabled'] is True
        assert result[0]['type'] == 'Cron'
        assert result[0]['times'] == '08:00'
        assert result[0]['timezone'] == 'UTC'
        assert 'Monday' in result[0]['weekdays']
        assert result[0]['start_date'] == '2024-01-01'
        assert result[0]['end_date'] == '2099-12-31'

    def test_disabled_schedule_parsed(self, schedule_part_disabled):
        result = ext.parse_schedule(schedule_part_disabled)
        assert len(result) == 1
        assert result[0]['enabled'] is False
        assert result[0]['timezone'] == 'GMT Standard Time'

    def test_multiple_schedules(self, schedule_part_multiple):
        result = ext.parse_schedule(schedule_part_multiple)
        assert len(result) == 2
        enabled_schedules = [s for s in result if s['enabled']]
        assert len(enabled_schedules) == 1
        assert '06:00' in enabled_schedules[0]['times']
        assert '18:00' in enabled_schedules[0]['times']

    def test_empty_parts_returns_empty(self):
        assert ext.parse_schedule([]) == []

    def test_no_schedule_part_returns_empty(self):
        parts = [{'path': 'pipeline.json', 'payload': 'e30='}]
        assert ext.parse_schedule(parts) == []

    def test_schedule_without_dates(self):
        payload = base64.b64encode(json.dumps({
            'schedules': [{
                'enabled': True,
                'configuration': {'type': 'Cron', 'times': ['09:00']}
            }]
        }).encode()).decode()
        parts = [{'path': '.schedules', 'payload': payload}]
        result = ext.parse_schedule(parts)
        assert result[0]['start_date'] == ''
        assert result[0]['end_date'] == ''

    def test_invalid_base64_returns_empty(self):
        parts = [{'path': '.schedules', 'payload': 'not-valid-base64!!!'}]
        assert ext.parse_schedule(parts) == []

    def test_interval_schedule(self):
        payload = base64.b64encode(json.dumps({
            'schedules': [{
                'enabled': True,
                'configuration': {'type': 'Interval', 'interval': 30}
            }]
        }).encode()).decode()
        parts = [{'path': '.schedules', 'payload': payload}]
        result = ext.parse_schedule(parts)
        assert result[0]['interval'] == '30'


# ── get_invoke_targets ────────────────────────────────────────────────────────

class TestGetInvokeTargets:
    def test_execute_pipeline_found(self, pipeline_part_with_invoke):
        result = ext.get_invoke_targets(pipeline_part_with_invoke)
        assert 'pl_child_pipeline' in result

    def test_nested_invoke_pipeline(self, pipeline_part_nested_invoke):
        result = ext.get_invoke_targets(pipeline_part_nested_invoke)
        assert 'pl_bronze_load' in result

    def test_no_invokes_returns_empty(self):
        import base64, json
        payload = base64.b64encode(json.dumps({
            'properties': {'activities': [
                {'name': 'Copy data', 'type': 'Copy', 'typeProperties': {}}
            ]}
        }).encode()).decode()
        parts = [{'path': 'pipeline.json', 'payload': payload}]
        assert ext.get_invoke_targets(parts) == []

    def test_empty_parts(self):
        assert ext.get_invoke_targets([]) == []

    def test_deduplicates_same_child(self):
        import base64, json
        payload = base64.b64encode(json.dumps({
            'properties': {'activities': [
                {'name': 'A', 'type': 'ExecutePipeline',
                 'typeProperties': {'pipeline': {'referenceName': 'pl_same'}}},
                {'name': 'B', 'type': 'ExecutePipeline',
                 'typeProperties': {'pipeline': {'referenceName': 'pl_same'}}},
            ]}
        }).encode()).decode()
        parts = [{'path': 'pipeline.json', 'payload': payload}]
        result = ext.get_invoke_targets(parts)
        assert result.count('pl_same') == 1

    def test_skips_schedules_and_platform_paths(self):
        import base64
        parts = [
            {'path': '.schedules', 'payload': base64.b64encode(b'{}').decode()},
            {'path': 'activity.platform', 'payload': base64.b64encode(b'{}').decode()},
        ]
        assert ext.get_invoke_targets(parts) == []


# ── get_definition_parts (mocked HTTP) ───────────────────────────────────────

class TestGetDefinitionParts:
    def setup_method(self):
        ext.WORKSPACE_ID = 'ws-test-id'
        ext.FAB = {'Authorization': 'Bearer test-token'}

    @pytest.mark.integration
    def test_200_response_returns_parts(self):
        mock_parts = [{'path': '.schedules', 'payload': 'abc'}]
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {'definition': {'parts': mock_parts}}

        with patch('requests.post', return_value=mock_resp):
            result = ext.get_definition_parts('item-123')

        assert result == mock_parts

    @pytest.mark.integration
    def test_202_async_returns_parts_after_poll(self):
        mock_parts = [{'path': 'pipeline.json', 'payload': 'xyz'}]

        async_resp = MagicMock()
        async_resp.status_code = 202
        async_resp.headers = {'Location': 'https://api.fabric.microsoft.com/v1/operations/op-1'}

        poll_resp = MagicMock()
        poll_resp.status_code = 200
        poll_resp.json.return_value = {'definition': {'parts': mock_parts}}

        with patch('requests.post', return_value=async_resp), \
             patch('requests.get', return_value=poll_resp), \
             patch('time.sleep'):
            result = ext.get_definition_parts('item-456')

        assert result == mock_parts

    @pytest.mark.integration
    def test_202_missing_location_header_returns_empty(self):
        async_resp = MagicMock()
        async_resp.status_code = 202
        async_resp.headers = {}

        with patch('requests.post', return_value=async_resp):
            result = ext.get_definition_parts('item-789')

        assert result == []

    @pytest.mark.integration
    def test_error_response_returns_empty(self):
        mock_resp = MagicMock()
        mock_resp.status_code = 500

        with patch('requests.post', return_value=mock_resp):
            result = ext.get_definition_parts('item-err')

        assert result == [] or result is None

    @pytest.mark.integration
    def test_network_exception_returns_none(self):
        with patch('requests.post', side_effect=Exception('timeout')):
            result = ext.get_definition_parts('item-net')

        assert result is None


# ── get_last_run (mocked HTTP) ────────────────────────────────────────────────

class TestGetLastRun:
    def setup_method(self):
        ext.WORKSPACE_ID = 'ws-test-id'
        ext.FAB = {'Authorization': 'Bearer test-token'}

    @pytest.mark.integration
    def test_returns_last_run_time_and_status(self):
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {'value': [
            {'startTime': '2024-03-15T08:30:00Z', 'status': 'Succeeded'}
        ]}

        with patch('requests.get', return_value=mock_resp):
            run_time, status = ext.get_last_run('item-123')

        assert run_time == '2024-03-15 08:30 UTC'
        assert status == 'Succeeded'

    @pytest.mark.integration
    def test_empty_run_list_returns_never_run(self):
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {'value': []}

        with patch('requests.get', return_value=mock_resp):
            run_time, status = ext.get_last_run('item-456')

        assert run_time == 'Never run'
        assert status == 'N/A'

    @pytest.mark.integration
    def test_403_returns_no_access(self):
        mock_resp = MagicMock()
        mock_resp.status_code = 403

        with patch('requests.get', return_value=mock_resp):
            run_time, status = ext.get_last_run('item-789')

        assert run_time == 'No access'
        assert status == 'No access'

    @pytest.mark.integration
    def test_404_returns_never_run(self):
        mock_resp = MagicMock()
        mock_resp.status_code = 404

        with patch('requests.get', return_value=mock_resp):
            run_time, status = ext.get_last_run('item-404')

        assert run_time == 'Never run'
        assert status == 'N/A'

    @pytest.mark.integration
    def test_network_exception_returns_error(self):
        with patch('requests.get', side_effect=Exception('timeout')):
            run_time, status = ext.get_last_run('item-net')

        assert run_time == 'Error'
        assert status == 'Error'


# ── get_token ─────────────────────────────────────────────────────────────────

class TestGetToken:
    def test_silent_token_returned_if_cached(self):
        mock_app = MagicMock()
        mock_cache = MagicMock()
        mock_cache.has_state_changed = False
        mock_app.get_accounts.return_value = [{'username': 'test@example.com'}]
        mock_app.acquire_token_silent.return_value = {'access_token': 'cached-token'}

        result = ext.get_token(['scope'], mock_app, mock_cache)

        assert result == 'cached-token'
        mock_app.acquire_token_interactive.assert_not_called()

    def test_interactive_used_when_no_accounts(self):
        mock_app = MagicMock()
        mock_cache = MagicMock()
        mock_cache.has_state_changed = False
        mock_app.get_accounts.return_value = []
        mock_app.acquire_token_interactive.return_value = {'access_token': 'new-token'}

        result = ext.get_token(['scope'], mock_app, mock_cache)

        assert result == 'new-token'

    def test_device_flow_used_on_interactive_failure(self):
        mock_app = MagicMock()
        mock_cache = MagicMock()
        mock_cache.has_state_changed = False
        mock_app.get_accounts.return_value = []
        mock_app.acquire_token_interactive.side_effect = Exception('no browser')
        mock_app.initiate_device_flow.return_value = {'message': 'Go to https://aka.ms/devicelogin'}
        mock_app.acquire_token_by_device_flow.return_value = {'access_token': 'device-token'}

        result = ext.get_token(['scope'], mock_app, mock_cache)

        assert result == 'device-token'

    def test_sys_exit_on_failed_auth(self):
        mock_app = MagicMock()
        mock_cache = MagicMock()
        mock_app.get_accounts.return_value = []
        mock_app.acquire_token_interactive.return_value = {'error': 'invalid_grant'}

        with pytest.raises(SystemExit):
            ext.get_token(['scope'], mock_app, mock_cache)


# ── make_progress ─────────────────────────────────────────────────────────────

class TestMakeProgress:
    def test_returns_progress_instance(self):
        from rich.progress import Progress
        result = ext.make_progress()
        assert isinstance(result, Progress)


# ── get_token – additional branches ──────────────────────────────────────────

class TestGetTokenCachePaths:
    def test_device_flow_error_message_then_exit(self):
        mock_app = MagicMock()
        mock_cache = MagicMock()
        mock_cache.has_state_changed = False
        mock_app.get_accounts.return_value = []
        mock_app.acquire_token_interactive.side_effect = Exception('no browser')
        mock_app.initiate_device_flow.return_value = {
            'error': 'access_denied',
            'error_description': 'User cancelled',
        }

        with pytest.raises(SystemExit):
            ext.get_token(['scope'], mock_app, mock_cache)

    def test_cache_written_when_state_changed(self):
        mock_app = MagicMock()
        mock_cache = MagicMock()
        mock_cache.has_state_changed = True
        mock_cache.serialize.return_value = 'serialized-data'
        mock_app.get_accounts.return_value = [{'username': 'test@example.com'}]
        mock_app.acquire_token_silent.return_value = {'access_token': 'token'}

        mock_cf = MagicMock()
        with patch.object(ext, 'CACHE_FILE', mock_cf):
            ext.get_token(['scope'], mock_app, mock_cache)

        mock_cf.write_text.assert_called_once_with('serialized-data', encoding='utf-8')

    def test_chmod_failure_silently_ignored(self):
        mock_app = MagicMock()
        mock_cache = MagicMock()
        mock_cache.has_state_changed = True
        mock_cache.serialize.return_value = 'data'
        mock_app.get_accounts.return_value = [{'username': 'test@example.com'}]
        mock_app.acquire_token_silent.return_value = {'access_token': 'token'}

        mock_cf = MagicMock()
        mock_cf.chmod.side_effect = AttributeError
        with patch.object(ext, 'CACHE_FILE', mock_cf):
            result = ext.get_token(['scope'], mock_app, mock_cache)

        assert result == 'token'


# ── get_definition_parts – inner exception branch ────────────────────────────

class TestGetDefinitionPartsRetry:
    def setup_method(self):
        ext.WORKSPACE_ID = 'ws-test-id'
        ext.FAB = {'Authorization': 'Bearer test-token'}

    @pytest.mark.integration
    def test_inner_poll_exception_swallowed_retries_succeed(self):
        async_resp = MagicMock()
        async_resp.status_code = 202
        async_resp.headers = {'Location': 'https://api.example.com/ops/1'}

        mock_parts = [{'path': '.schedules', 'payload': 'abc'}]
        ok_resp = MagicMock()
        ok_resp.status_code = 200
        ok_resp.json.return_value = {'definition': {'parts': mock_parts}}

        with patch('requests.post', return_value=async_resp), \
             patch('requests.get', side_effect=[Exception('timeout'), ok_resp]), \
             patch('time.sleep'):
            result = ext.get_definition_parts('item-retry')

        assert result == mock_parts

    @pytest.mark.integration
    def test_polling_timeout_returns_empty(self):
        async_resp = MagicMock()
        async_resp.status_code = 202
        async_resp.headers = {'Location': 'https://api.example.com/ops/1'}

        never_done = MagicMock()
        never_done.status_code = 202

        with patch('requests.post', return_value=async_resp), \
             patch('requests.get', return_value=never_done), \
             patch('time.sleep'):
            result = ext.get_definition_parts('item-timeout')

        assert result == []


# ── get_invoke_targets – additional branches ──────────────────────────────────

class TestGetInvokeTargetsAdditional:
    def test_string_ref_used_as_name(self):
        payload = base64.b64encode(json.dumps({
            'properties': {'activities': [
                {'name': 'A', 'type': 'ExecutePipeline',
                 'typeProperties': {'pipeline': 'pl_string_name'}}
            ]}
        }).encode()).decode()
        parts = [{'path': 'pipeline.json', 'payload': payload}]
        result = ext.get_invoke_targets(parts)
        assert 'pl_string_name' in result

    def test_body_dict_activities_scanned(self):
        payload = base64.b64encode(json.dumps({
            'properties': {'activities': [
                {'name': 'ForEach', 'type': 'ForEach',
                 'typeProperties': {
                     'body': {'activities': [
                         {'name': 'Inner', 'type': 'ExecutePipeline',
                          'typeProperties': {'pipeline': {'referenceName': 'pl_inner_body'}}}
                     ]}
                 }}
            ]}
        }).encode()).decode()
        parts = [{'path': 'pipeline.json', 'payload': payload}]
        result = ext.get_invoke_targets(parts)
        assert 'pl_inner_body' in result

    def test_corrupt_payload_silently_skipped(self):
        parts = [{'path': 'pipeline.json', 'payload': 'NOT_VALID_BASE64!!'}]
        result = ext.get_invoke_targets(parts)
        assert result == []


# ── get_last_run – non-200/403/404 status ────────────────────────────────────

class TestGetLastRunAdditional:
    def setup_method(self):
        ext.WORKSPACE_ID = 'ws-test-id'
        ext.FAB = {'Authorization': 'Bearer test-token'}

    @pytest.mark.integration
    def test_500_response_returns_error(self):
        mock_resp = MagicMock()
        mock_resp.status_code = 500

        with patch('requests.get', return_value=mock_resp):
            run_time, status = ext.get_last_run('item-500')

        assert run_time == 'Error'
        assert status == 'Error'


# ── make_header ───────────────────────────────────────────────────────────────

class TestMakeHeader:
    def _ws(self):
        import openpyxl
        return openpyxl.Workbook().active

    def test_title_written_to_merged_row1(self):
        ws = self._ws()
        ext.make_header(ws, 'My Title', 'My Subtitle', ['Col A', 'Col B', 'Col C'])
        assert ws.cell(1, 1).value == 'My Title'

    def test_subtitle_written_to_row2(self):
        ws = self._ws()
        ext.make_header(ws, 'T', 'Sub', ['A', 'B'])
        assert ws.cell(2, 1).value == 'Sub'

    def test_column_headers_written_to_row3(self):
        ws = self._ws()
        cols = ['Name', 'Enabled', 'Type']
        ext.make_header(ws, 'T', 'S', cols)
        assert ws.cell(3, 1).value == 'Name'
        assert ws.cell(3, 2).value == 'Enabled'
        assert ws.cell(3, 3).value == 'Type'


# ── write_rows ────────────────────────────────────────────────────────────────

class TestWriteRows:
    def _ws(self):
        import openpyxl
        return openpyxl.Workbook().active

    def test_basic_value_written(self):
        ws = self._ws()
        ext.write_rows(ws, [{'Name': 'Pipeline A'}], ['Name'])
        assert ws.cell(4, 1).value == 'Pipeline A'

    def test_enabled_yes(self):
        ws = self._ws()
        ext.write_rows(ws, [{'Enabled': 'Yes'}], ['Enabled'])
        assert ws.cell(4, 1).value == 'Yes'

    def test_enabled_disabled_variant(self):
        ws = self._ws()
        ext.write_rows(ws, [{'Enabled': 'No (Disabled)'}], ['Enabled'])
        assert ws.cell(4, 1).value == 'No (Disabled)'

    def test_enabled_no_schedule(self):
        ws = self._ws()
        ext.write_rows(ws, [{'Enabled': 'No schedule'}], ['Enabled'])
        assert ws.cell(4, 1).value == 'No schedule'

    def test_status_succeeded(self):
        ws = self._ws()
        ext.write_rows(ws, [{'Last Run Status': 'Succeeded'}], ['Last Run Status'])
        assert ws.cell(4, 1).value == 'Succeeded'

    def test_status_completed(self):
        ws = self._ws()
        ext.write_rows(ws, [{'Last Run Status': 'Completed'}], ['Last Run Status'])
        assert ws.cell(4, 1).value == 'Completed'

    def test_status_failed(self):
        ws = self._ws()
        ext.write_rows(ws, [{'Last Run Status': 'Failed'}], ['Last Run Status'])
        assert ws.cell(4, 1).value == 'Failed'

    def test_status_inprogress(self):
        ws = self._ws()
        ext.write_rows(ws, [{'Last Run Status': 'InProgress'}], ['Last Run Status'])
        assert ws.cell(4, 1).value == 'InProgress'

    def test_status_other(self):
        ws = self._ws()
        ext.write_rows(ws, [{'Last Run Status': 'N/A'}], ['Last Run Status'])
        assert ws.cell(4, 1).value == 'N/A'

    def test_schedule_until_expired(self):
        ws = self._ws()
        ext.write_rows(ws, [{'Schedule Until': '2020-01-01'}], ['Schedule Until'])
        assert ws.cell(4, 1).value == '2020-01-01'

    def test_schedule_until_future(self):
        ws = self._ws()
        ext.write_rows(ws, [{'Schedule Until': '2099-12-31'}], ['Schedule Until'])
        assert ws.cell(4, 1).value == '2099-12-31'

    def test_schedule_until_empty(self):
        ws = self._ws()
        ext.write_rows(ws, [{'Schedule Until': ''}], ['Schedule Until'])
        assert ws.cell(4, 1).value == ''

    def test_multiple_rows_alternate_fill(self):
        ws = self._ws()
        rows = [{'Name': 'A'}, {'Name': 'B'}, {'Name': 'C'}]
        ext.write_rows(ws, rows, ['Name'])
        assert ws.cell(4, 1).value == 'A'
        assert ws.cell(5, 1).value == 'B'
        assert ws.cell(6, 1).value == 'C'


# ── set_widths ────────────────────────────────────────────────────────────────

class TestSetWidths:
    def test_column_widths_applied(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ext.set_widths(ws, [20, 30, 15])
        assert ws.column_dimensions['A'].width == 20
        assert ws.column_dimensions['B'].width == 30
        assert ws.column_dimensions['C'].width == 15


# ── process_definition_items ──────────────────────────────────────────────────

class TestProcessDefinitionItems:
    def setup_method(self):
        ext.WORKSPACE_ID = 'ws-test-id'
        ext.FAB = {'Authorization': 'Bearer test-token'}
        ext.INCLUDE_UNSCHEDULED = True

    def _mock_progress(self):
        cm = MagicMock()
        cm.__enter__ = MagicMock(return_value=MagicMock())
        cm.__exit__ = MagicMock(return_value=False)
        return cm

    @pytest.mark.integration
    def test_returns_empty_when_no_items(self):
        resp = MagicMock()
        resp.json.return_value = {'value': []}

        with patch('requests.get', return_value=resp):
            result = ext.process_definition_items('DataPipeline', 'Data Pipeline')

        assert result == []

    @pytest.mark.integration
    def test_scheduled_item_returns_row(self):
        resp = MagicMock()
        resp.json.return_value = {'value': [{'displayName': 'pl_main', 'id': 'item-1'}]}

        schedule = [{'enabled': True, 'type': 'Cron', 'times': '08:00',
                     'timezone': 'UTC', 'weekdays': 'Monday', 'interval': '',
                     'start_date': '2024-01-01', 'end_date': '2099-12-31'}]

        with patch('requests.get', return_value=resp), \
             patch.object(ext, 'get_definition_parts', return_value=[]), \
             patch.object(ext, 'parse_schedule', return_value=schedule), \
             patch.object(ext, 'get_last_run', return_value=('2024-03-01 08:00 UTC', 'Succeeded')), \
             patch.object(ext, 'make_progress', return_value=self._mock_progress()):
            result = ext.process_definition_items('DataPipeline', 'Data Pipeline')

        assert len(result) == 1
        assert result[0]['Name'] == 'pl_main'
        assert result[0]['Enabled'] == 'Yes'

    @pytest.mark.integration
    def test_disabled_schedule_tag_reported(self):
        resp = MagicMock()
        resp.json.return_value = {'value': [{'displayName': 'pl_mixed', 'id': 'item-2'}]}

        schedules = [
            {'enabled': True, 'type': 'Cron', 'times': '08:00', 'timezone': 'UTC',
             'weekdays': 'Monday', 'interval': '', 'start_date': '', 'end_date': ''},
            {'enabled': False, 'type': 'Cron', 'times': '18:00', 'timezone': 'UTC',
             'weekdays': 'Monday', 'interval': '', 'start_date': '', 'end_date': ''},
        ]

        with patch('requests.get', return_value=resp), \
             patch.object(ext, 'get_definition_parts', return_value=[]), \
             patch.object(ext, 'parse_schedule', return_value=schedules), \
             patch.object(ext, 'get_last_run', return_value=('Never run', 'N/A')), \
             patch.object(ext, 'make_progress', return_value=self._mock_progress()):
            result = ext.process_definition_items('DataPipeline', 'Data Pipeline')

        assert len(result) == 2
        enabled = [r for r in result if r['Enabled'] == 'Yes']
        disabled = [r for r in result if 'Disabled' in r['Enabled']]
        assert len(enabled) == 1
        assert len(disabled) == 1

    @pytest.mark.integration
    def test_unscheduled_included_when_flag_true(self):
        ext.INCLUDE_UNSCHEDULED = True
        resp = MagicMock()
        resp.json.return_value = {'value': [{'displayName': 'pl_none', 'id': 'item-3'}]}

        with patch('requests.get', return_value=resp), \
             patch.object(ext, 'get_definition_parts', return_value=[]), \
             patch.object(ext, 'parse_schedule', return_value=[]), \
             patch.object(ext, 'get_last_run', return_value=('Never run', 'N/A')), \
             patch.object(ext, 'make_progress', return_value=self._mock_progress()):
            result = ext.process_definition_items('DataPipeline', 'Data Pipeline')

        assert len(result) == 1
        assert result[0]['Enabled'] == 'No schedule'

    @pytest.mark.integration
    def test_unscheduled_excluded_when_flag_false(self):
        ext.INCLUDE_UNSCHEDULED = False
        resp = MagicMock()
        resp.json.return_value = {'value': [{'displayName': 'pl_none', 'id': 'item-4'}]}

        with patch('requests.get', return_value=resp), \
             patch.object(ext, 'get_definition_parts', return_value=[]), \
             patch.object(ext, 'parse_schedule', return_value=[]), \
             patch.object(ext, 'get_last_run', return_value=('Never run', 'N/A')), \
             patch.object(ext, 'make_progress', return_value=self._mock_progress()):
            result = ext.process_definition_items('DataPipeline', 'Data Pipeline')

        assert result == []
